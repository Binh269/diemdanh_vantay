import base64
import ctypes
from ctypes import wintypes
import datetime
import hashlib
from mailbox import mbox
import os
import random
import shutil
import sqlite3
import string
import tempfile
import time
from tkinter import Image
import urllib.request as url_request
from django.conf import settings
from django.http import HttpResponse, HttpResponseServerError, JsonResponse
from django.shortcuts import render, redirect
from django.views import View
import xlwt
import xlrd
from openpyxl import load_workbook
from Show.traning import train_new_student
from .models import Model_Phong
from django.contrib import messages
import json
import numpy as np
from django.core.files.base import ContentFile
import torch
import cv2
from facenet_pytorch import InceptionResnetV1, MTCNN
from torchvision import transforms
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from scipy.spatial.distance import cosine
import pyodbc
from datetime import datetime, timedelta
from unidecode import unidecode
from collections import Counter
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
import openpyxl
from django.core.files.storage import default_storage
from collections import Counter, defaultdict
from django.utils import timezone
import pandas as pd
from pyfingerprint.pyfingerprint import PyFingerprint
import logging
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponseForbidden
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.auth.models import AbstractUser
from django.db import models
import mediapipe as mp
import math

def is_superuser_sql(user_id):
    conn = connect_to_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT is_superuser FROM auth_user WHERE id = ?", (user_id,))
        result = cursor.fetchone()
        if result and result[0] >= 1: 
            return True
        return False
    except Exception as e:
        print(f"Lỗi khi kiểm tra is_superuser: {e}")
        return False
    finally:
        cursor.close()
        conn.close()

def superuser(user_id):
    conn = connect_to_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT is_superuser FROM auth_user WHERE id = ?", (user_id,))
        result = cursor.fetchone()
        if result and result[0] == 2: 
            return True
        return False
    except Exception as e:
        print(f"Lỗi khi kiểm tra is_superuser: {e}")
        return False
    finally:
        cursor.close()
        conn.close()

def admin_required_sql(view_func):
    @login_required(login_url="login")
    def wrapper(request, *args, **kwargs):
        if not is_superuser_sql(request.user.id):
            return HttpResponseForbidden("Bạn không có quyền truy cập!")
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required(login_url="login")  
def index(request):
    return render(request, "index.html")

@login_required(login_url="login")
def Lich(request):
    if not is_superuser_sql(request.user.id):
        return render(request, "lich.html", {'restrict_class_selection': True})
    return render(request, "lich.html", {'restrict_class_selection': False})

@login_required(login_url="login")
def ql_sv(request):
    return render(request, "quanly.html")

@login_required(login_url="login")
@admin_required_sql
def Diemdanh(request):
    return render(request, 'diemdanh.html')

@login_required(login_url="login")
def thongke_view(request):
    return render(request, "thongke.html")

@login_required(login_url="login")
@admin_required_sql
def setting(request):
    return render(request, "setting.html")

def connect_to_db():
    return sqlite3.connect("db.sqlite3")


# device = torch.device("cuda")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = InceptionResnetV1(pretrained='vggface2').eval().to(device)
mp_face_detection = mp.solutions.face_detection
face_detection = mp_face_detection.FaceDetection(
    model_selection=0, 
    min_detection_confidence=0.5  
)
transform = transforms.Compose([
    transforms.ToPILImage(),
    transforms.Resize((128, 128)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
])

def embedding_to_string(embedding):
    return json.dumps(embedding.tolist())

def string_to_embedding(embedding_str):
    if embedding_str is None:
        raise ValueError("Embedding không được phép là None")
    return np.array(json.loads(embedding_str))

def preprocess_frame(frame):
    max_width = 320
    h, w = frame.shape[:2]
    if w > max_width:
        scale = max_width / w
        new_h, new_w = int(h * scale), int(w * scale)
        frame = cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA)
    return frame
embedding_cache = {}

def detect_face(frame, lophp, request):
    frame = preprocess_frame(frame)
    
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    
    results = face_detection.process(frame_rgb)
    detected_labels = []

    if not results.detections:  
        return detected_labels

    h, w = frame.shape[:2]
    orig_h, orig_w = frame.shape[:2]
    scale_x, scale_y = orig_w / w, orig_h / h

    if lophp not in embedding_cache:
        conn = connect_to_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT thongtin.mssv, thongtin.khuonmat 
            FROM thongtin 
            INNER JOIN dkmh ON dkmh.mssv = thongtin.mssv 
            WHERE dkmh.lophp = ? AND thongtin.khuonmat IS NOT NULL
        """, (lophp,))
        embeddings_data = cursor.fetchall()
        conn.close()

        embedding_cache[lophp] = {
            'embeddings': np.array([string_to_embedding(row[1]) for row in embeddings_data if row[1] is not None]),
            'labels': [row[0] for row in embeddings_data if row[1] is not None]
        }

    dataset_embeddings = embedding_cache[lophp]['embeddings']
    dataset_labels = embedding_cache[lophp]['labels']

    if not dataset_embeddings.size:
        return detected_labels

    settings = get_user_settings(request, as_dict=True)
    MIN_AREA_THRESHOLD = settings['kichthuoc'] *(0.5)
    MIN_DISTANCE_THRESHOLD = settings['nguong']

    for detection in results.detections:
        bbox = detection.location_data.relative_bounding_box
        x1 = int(bbox.xmin * w * scale_x)
        y1 = int(bbox.ymin * h * scale_y)
        x2 = int((bbox.xmin + bbox.width) * w * scale_x)
        y2 = int((bbox.ymin + bbox.height) * h * scale_y)
        
        width = x2 - x1
        height = y2 - y1
        area = width * height

        if area < MIN_AREA_THRESHOLD:
            continue

        face = frame[int(y1 / scale_y):int(y2 / scale_y), int(x1 / scale_x):int(x2 / scale_x)]
        if face.size == 0:
            continue

        face_tensor = transform(face).unsqueeze(0).to(device)
        face_embedding = model(face_tensor).detach().cpu().numpy().flatten()

        distances = 1 - np.dot(dataset_embeddings, face_embedding) / (
            np.linalg.norm(dataset_embeddings, axis=1) * np.linalg.norm(face_embedding)
        )
        min_distance_idx = np.argmin(distances)
        min_distance = distances[min_distance_idx]
        label = dataset_labels[min_distance_idx] if min_distance < MIN_DISTANCE_THRESHOLD else "Không biết"

        detected_labels.append({
            'label': label,
            'box': [int(x1), int(y1), int(x2), int(y2)]
        })

    return detected_labels


@login_required(login_url="login")
def get_user_settings(request, as_dict=False):
    conn = connect_to_db()
    cursor = conn.cursor()
    try:
        magv = str(request.user.id)
        cursor.execute("SELECT kichthuoc, nguong FROM setting WHERE magv = ?", (magv,))
        result = cursor.fetchone()

        if result:
            kichthuoc = int(result[0]) if result[0] is not None else 12000
            nguong = float(result[1]) if result[1] is not None else 0.2
            data = {'kichthuoc': kichthuoc, 'nguong': nguong}
        else:
            data = {'kichthuoc': 12000, 'nguong': 0.2}

        print(f"DEBUG: User settings for magv {magv}: {data}")
        if as_dict:
            return data
        return JsonResponse(data)

    except Exception as e:
        print(f"Lỗi khi lấy cài đặt người dùng: {e}")
        data = {'kichthuoc': 12000, 'nguong': 0.2}
        if as_dict:
            return data
        return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'}, status=500)

    finally:
        cursor.close()
        conn.close()

@login_required(login_url="login")
@admin_required_sql
def save_settings(request):
    if request.method != "POST":
        return JsonResponse({'status': 'fail', 'message': 'Phương thức không hợp lệ'}, status=405)

    conn = connect_to_db()
    cursor = conn.cursor()
    
    try:
        magv = str(request.user.id)
        kichthuoc = request.POST.get('kichthuoc')
        nguong = request.POST.get('nguong')

        if not kichthuoc or not nguong:
            return JsonResponse({'status': 'fail', 'message': 'Vui lòng nhập đầy đủ kích thước và ngưỡng'})

        kichthuoc = int(kichthuoc)
        nguong = float(nguong)

        cursor.execute("SELECT COUNT(*) FROM setting WHERE magv = ?", (magv,))
        exists = cursor.fetchone()[0] > 0

        if exists:
            cursor.execute("UPDATE setting SET kichthuoc = ?, nguong = ? WHERE magv = ?", 
                         (kichthuoc, nguong, magv))
        else:
            cursor.execute("INSERT INTO setting (magv, kichthuoc, nguong) VALUES (?, ?, ?)",
                         (magv, kichthuoc, nguong))

        conn.commit()
        return JsonResponse({'status': 'success', 'message': 'Lưu cài đặt thành công'})

    except ValueError as e:
        return JsonResponse({'status': 'fail', 'message': f'Dữ liệu không hợp lệ: {str(e)}'})
    except Exception as e:
        print(f"Lỗi khi lưu cài đặt: {e}")
        return JsonResponse({'status': 'fail', 'message': f'Lỗi hệ thống: {str(e)}'})

    finally:
        cursor.close()
        conn.close()

@login_required(login_url="login")
def settings_view(request):
    conn = connect_to_db()
    cursor = conn.cursor()
    try:
        user_id = request.user.id  
        has_permission = superuser(user_id) 

        if request.method == "GET":
            if has_permission:
                cursor.execute("SELECT id, username, is_superuser, email, last_name FROM auth_user")
                users = [{'id': row[0], 'username': row[1], 'is_superuser': row[2], 'email': row[3], 'name': row[4]} for row in cursor.fetchall()]
            else:
                users = []  

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'users': users, 'has_permission': has_permission})
            else:
                data = {'users': users, 'has_permission': has_permission, 'current_magv': str(user_id)}
                return render(request, 'settings.html', data)

        elif request.method == "POST":
            if not has_permission:
                return JsonResponse({'status': 'fail', 'message': 'Không đủ quyền để thực hiện hành động này'}, status=403)

            action = request.POST.get('action')
            if action == 'add':
                username = request.POST.get('username')
                password = request.POST.get('password')
                name = request.POST.get('name')
                email = request.POST.get('email')
                is_superuser = request.POST.get('is_superuser')
                date = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

                if not username or not password:
                    return JsonResponse({'status': 'fail', 'message': 'Thiếu username hoặc password'})

                cursor.execute("SELECT COUNT(*) FROM auth_user WHERE username = ?", (username,))
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'fail', 'message': 'Username đã tồn tại'})

                cursor.execute("SELECT COUNT(*) FROM auth_user WHERE email = ?", (email,))
                if cursor.fetchone()[0] > 0:
                    return JsonResponse({'status': 'fail', 'message': 'Gmail đã tồn tại'})

                hashed_password = make_password(password)
                cursor.execute(
                    "INSERT INTO auth_user (username, password, last_name, email, is_superuser, is_staff, is_active, date_joined) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (username, hashed_password, name, email, is_superuser, 1 if is_superuser else 0, 1, date)
                )
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Thêm tài khoản thành công'})

            elif action == 'update':
                user_id_to_update = request.POST.get('user_id')
                is_superuser = request.POST.get('is_superuser')
                username = request.POST.get('username')
                name = request.POST.get('name')
                email = request.POST.get('email')
                password = request.POST.get('password')

                if not user_id_to_update:
                    return JsonResponse({'status': 'fail', 'message': 'Thiếu user_id'})

                if username:
                    cursor.execute("SELECT COUNT(*) FROM auth_user WHERE username = ? AND id != ?", (username, user_id_to_update))
                    if cursor.fetchone()[0] > 0:
                        return JsonResponse({'status': 'fail', 'message': 'Username đã tồn tại'})

                if email:
                    cursor.execute("SELECT COUNT(*) FROM auth_user WHERE email = ? AND id != ?", (email, user_id_to_update))
                    if cursor.fetchone()[0] > 0:
                        return JsonResponse({'status': 'fail', 'message': 'Email đã tồn tại'})

                if password:
                    hashed_password = make_password(password)
                    cursor.execute(
                        "UPDATE auth_user SET username = ?, last_name = ?, email = ?, is_superuser = ?, password = ? WHERE id = ?",
                        (username or None, name or None, email or None,is_superuser, hashed_password, user_id_to_update)
                    )
                else:
                    cursor.execute(
                        "UPDATE auth_user SET username = ?, last_name = ?, email = ?, is_superuser = ? WHERE id = ?",
                        (username or None, name or None, email or None, is_superuser , user_id_to_update)
                    )

                if cursor.rowcount == 0:
                    return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy người dùng để cập nhật'})
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Cập nhật tài khoản thành công'})

            elif action == 'delete':
                user_id_to_delete = request.POST.get('user_id')
                if not user_id_to_delete:
                    return JsonResponse({'status': 'fail', 'message': 'Thiếu user_id'})
                if user_id_to_delete == str(user_id): 
                    return JsonResponse({'status': 'fail', 'message': 'Không thể xóa tài khoản đang đăng nhập'})

                cursor.execute("DELETE FROM auth_user WHERE id = ?", (user_id_to_delete,))
                if cursor.rowcount == 0:
                    return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy người dùng để xóa'})
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Xóa tài khoản thành công'})

            return JsonResponse({'status': 'fail', 'message': 'Hành động không hợp lệ'})

        else:
            return JsonResponse({'status': 'fail', 'message': 'Phương thức không hợp lệ'}, status=405)

    except Exception as e:
        print(f"Error in settings_view: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'fail', 'message': f'Error: {str(e)}'})
        return HttpResponseServerError(f'Error: {str(e)}')
    finally:
        cursor.close()
        conn.close()

@login_required(login_url="login")
def get_available_classes(request):
    conn = connect_to_db()
    cursor = conn.cursor()

    try:
        magv = str(request.user.id)
        current_date = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT DISTINCT lophp, tiet, magv 
            FROM lich 
            WHERE ngay = ? AND magv = ?
            ORDER BY lophp, tiet
        """, (current_date, magv))
        schedule = cursor.fetchall()

        available_classes = list(set(row[0] for row in schedule))
        available_periods = list(set(row[1] for row in schedule))

        data = {
            'status': 'success',
            'classes': available_classes,
            'periods': available_periods,
            'magv': [magv], 
            'schedule': [{'lophp': row[0], 'tiet': row[1], 'magv': row[2]} for row in schedule]
        }
        print(f"DEBUG: Available classes for magv {magv}: {data}") 
        return JsonResponse(data)

    except Exception as e:
        print(f"Lỗi khi lấy lịch học: {e}")
        return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})

    finally:
        cursor.close()
        conn.close()

@csrf_exempt
def diemdanh(request):
    if request.method == "POST":
        try:
            image_data = request.POST.get('image')
            lophp = request.POST.get('lophp')
            tiet = request.POST.get('tiet')

            if not image_data:
                return JsonResponse({'success': False, 'message': 'Không nhận được ảnh!'})
            if not lophp or not tiet:
                return JsonResponse({'success': False, 'message': 'Vui lòng chọn lớp và tiết!'})

            magv = str(request.user.id)

            format, imgstr = image_data.split(';base64,')
            ext = format.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr), name=f'temp.{ext}')

            nparr = np.frombuffer(data.read(), np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

            detected_labels = detect_face(frame, lophp, request)

            if detected_labels:
                for detection in detected_labels:
                    label = detection['label']
                    if label != "Unknown":
                        mssv = label.split('_')[0]
                        conn = connect_to_db()
                        cursor = conn.cursor()
                        
                        current_date = datetime.now().strftime('%Y-%m-%d')
                        cursor.execute("""
                            SELECT malich 
                            FROM lich 
                            WHERE lophp = ? AND tiet = ? AND ngay = ? AND magv = ?
                        """, (lophp, tiet, current_date, magv))
                        lich_row = cursor.fetchone()
                        
                        if not lich_row:
                            conn.close()
                            return JsonResponse({'success': False, 'message': 'Không tìm thấy lịch học phù hợp hoặc bạn không có quyền điểm danh lớp này!'})
                        
                        malich = lich_row[0]
                        print("Mã lịch:", malich)

                        # Kiểm tra xem sinh viên đã điểm danh trong 1 phút gần nhất chưa
                        one_minute_ago = (datetime.now() - timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M:%S')
                        cursor.execute("""
                            SELECT time 
                            FROM diemdanh 
                            WHERE mssv = ? AND malich = ? AND time >= ?
                        """, (mssv, malich, one_minute_ago))
                        recent_diemdanh = cursor.fetchone()

                        if recent_diemdanh:
                            cursor.execute("SELECT name FROM thongtin WHERE mssv = ?", (mssv,))
                            student = cursor.fetchone()
                            conn.close()
                            return JsonResponse({
                                'success': False,
                                'mssv': mssv,
                                'name': student[0] if student else 'Unknown',
                                'labels': detected_labels,
                                'message': 'Sinh viên đã được điểm danh trong 1 phút gần nhất!'
                            })

                        # Nếu chưa điểm danh, tiếp tục xử lý
                        cursor.execute("SELECT name FROM thongtin WHERE mssv = ?", (mssv,))
                        student = cursor.fetchone()
                        
                        if student:
                            name = student[0]
                            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            
                            cursor.execute("""
                                INSERT INTO diemdanh (mssv, time, malich)
                                VALUES (?, ?, ?)
                            """, (mssv, current_time, malich))
                            conn.commit()
                            
                            conn.close()
                            return JsonResponse({
                                'success': True,
                                'mssv': mssv,
                                'name': name,
                                'lophp': lophp,
                                'tiet': tiet,
                                'magv': magv,  
                                'time': current_time,
                                'labels': detected_labels,
                                'message': 'Đã điểm danh'
                            })
                        
                        conn.close()
            return JsonResponse({
                'success': False,
                'labels': detected_labels,
                'message': 'Không nhận diện được khuôn mặt'
            })
            
        except Exception as e:
            print(f"Lỗi: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

@login_required(login_url="login")
def get_attendance_history(request):
    lophp = request.GET.get('lophp')
    tiet = request.GET.get('tiet')
    magv = str(request.user.id)

    if not lophp or not tiet:
        return JsonResponse({'status': 'fail', 'message': 'Vui lòng cung cấp lớp và tiết'})

    conn = connect_to_db()
    cursor = conn.cursor()

    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT malich 
            FROM lich 
            WHERE lophp = ? AND tiet = ? AND ngay = ? AND magv = ?
        """, (lophp, tiet, current_date, magv))
        lich_row = cursor.fetchone()

        if not lich_row:
            return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy lịch học phù hợp'})

        malich = lich_row[0]

        cursor.execute("""
            SELECT diemdanh.mssv, thongtin.name, diemdanh.time 
            FROM diemdanh 
            INNER JOIN thongtin ON diemdanh.mssv = thongtin.mssv 
            WHERE diemdanh.malich = ?
            ORDER BY diemdanh.time DESC
        """, (malich,))
        history = cursor.fetchall()

        data = {
            'status': 'success',
            'history': [
                {
                    'mssv': row[0],
                    'name': row[1]
                } for row in history
            ]
        }
        return JsonResponse(data)

    except Exception as e:
        print(f"Lỗi khi lấy lịch sử điểm danh: {e}")
        return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})

    finally:
        cursor.close()
        conn.close()

def vietnamese_last_word_sort_key(name):
    words = name.strip().split()
    last_word = words[-1] if words else "" 
    
    last_word_no_accent = unidecode(last_word).lower()
    
    vietnamese_order = (
        'a', 'ă', 'â', 'b', 'c', 'd', 'đ', 'e', 'ê', 'g', 'h', 'i', 
        'k', 'l', 'm', 'n', 'o', 'ô', 'ơ', 'p', 'q', 'r', 's', 't', 
        'u', 'ư', 'v', 'x', 'y'
    )
    
    first_char = last_word_no_accent[0] if last_word_no_accent else ''
    sort_index = vietnamese_order.index(first_char) if first_char in vietnamese_order else float('inf')
    return (sort_index, last_word_no_accent, unidecode(name).lower())  

@login_required(login_url="login")
def get_class_members(request):
    lophp = request.GET.get('lophp')
    tiet = request.GET.get('tiet')
    magv = str(request.user.id)
    if not lophp or not tiet:
        return JsonResponse({'status': 'fail', 'message': 'Vui lòng cung cấp lớp và tiết'})

    conn = connect_to_db()
    cursor = conn.cursor()

    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT malich 
            FROM lich 
            WHERE lophp = ? AND tiet = ? AND ngay = ? AND magv = ?
        """, (lophp, tiet, current_date,magv))
        lich_row = cursor.fetchone()

        malich = lich_row[0] if lich_row else None

        cursor.execute("""
            SELECT DISTINCT thongtin.mssv, thongtin.name, thongtin.lop 
            FROM thongtin 
            INNER JOIN dkmh ON dkmh.mssv = thongtin.mssv 
            WHERE dkmh.lophp = ?
        """, (lophp,))
        students = cursor.fetchall()

        students_sorted = sorted(students, key=lambda x: vietnamese_last_word_sort_key(x[1]))

        attended_students = set()
        if malich:
            cursor.execute("""
                SELECT mssv 
                FROM diemdanh 
                WHERE malich = ?
            """, (malich,))
            attended_students = set(row[0] for row in cursor.fetchall())

        data = {
            'members': [
                {
                    'stt': idx + 1, 
                    'mssv': student[0],
                    'name': student[1],  
                    'lophp': student[2],
                    'attended': student[0] in attended_students
                }
                for idx, student in enumerate(students_sorted)
            ]
        }
        return JsonResponse(data)

    except Exception as e:
        print(f"Lỗi khi truy vấn dữ liệu thành viên lớp: {e}")
        return JsonResponse({'status': 'fail', 'message': 'Lỗi khi lấy dữ liệu thành viên lớp'})

    finally:
        cursor.close()
        conn.close()
@csrf_exempt
def manual_attendance(request):
    if request.method == "POST":
        try:
            mssv = request.POST.get('mssv')
            lophp = request.POST.get('lophp')
            tiet = request.POST.get('tiet')
            attended = request.POST.get('attended') 
            magv = str(request.user.id)
            if not mssv or not lophp or not tiet:
                return JsonResponse({'success': False, 'message': 'Thiếu thông tin cần thiết!'})

            conn = connect_to_db()
            cursor = conn.cursor()

            current_date = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("""
                SELECT malich 
                FROM lich 
                WHERE lophp = ? AND tiet = ? AND ngay = ? And magv = ?
            """, (lophp, tiet, current_date,magv))
            lich_row = cursor.fetchone()

            if not lich_row:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không tìm thấy lịch học phù hợp!'})
            
            malich = lich_row[0]

            cursor.execute("SELECT name FROM thongtin WHERE mssv = ?", (mssv,))
            student = cursor.fetchone()
            if not student:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không tìm thấy sinh viên!'})

            if (attended =="true"):
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    INSERT INTO diemdanh (mssv, time, malich)
                    VALUES (?, ?, ?)
                """, (mssv, current_time, malich))
            else:
                cursor.execute("""
                    DELETE FROM diemdanh 
                    WHERE mssv = ? AND malich = ?
                """, (mssv, malich))

            conn.commit()
            conn.close()
            return JsonResponse({'success': True, 'message': 'Cập nhật điểm danh thành công'})

        except Exception as e:
            print(f"Lỗi khi cập nhật điểm danh thủ công: {e}")
            return JsonResponse({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ'})

DLL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "models", "ftrScanAPI.dll")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
logger.addHandler(logging.StreamHandler())

# Định nghĩa các constants theo Futronic SDK
FTR_RETCODE_OK = 0
FTR_RETCODE_INVALID_ARG = 1
FTR_RETCODE_ALREADY_IN_USE = 2
FTR_RETCODE_NO_MEMORY = 3
FTR_RETCODE_FRAME_SOURCE_NOT_SET = 4
FTR_RETCODE_CANCELED_BY_USER = 5
FTR_RETCODE_INTERNAL_ERROR = 6
FTR_RETCODE_DEVICE_NOT_CONNECTED = 7
FTR_RETCODE_DEVICE_FAILURE = 8
FTR_RETCODE_FAKE_SOURCE = 9
FTR_RETCODE_INVALID_PURPOSE = 10

# Frame source constants
FSD_FUTRONIC_USB = 1
FSD_UNDEFINED = 0

# Parameters
FTR_PARAM_IMAGE_WIDTH = 1
FTR_PARAM_IMAGE_HEIGHT = 2
FTR_PARAM_IMAGE_SIZE = 3
FTR_PARAM_CB_FRAME_SOURCE = 4
FTR_PARAM_CB_CONTROL = 5
FTR_PARAM_MAX_TEMPLATE_SIZE = 6
FTR_PARAM_MAX_FAR_REQUESTED = 7
FTR_PARAM_MAX_FARN_REQUESTED = 8
FTR_PARAM_SYS_ERROR_CODE = 9
FTR_PARAM_FAKE_DETECT = 10
FTR_PARAM_FFD_CONTROL = 11
FTR_PARAM_MAX_MODELS = 12

# Purpose constants
FTR_PURPOSE_ENROLL = 1
FTR_PURPOSE_IDENTIFY = 2

# Signal constants
FTR_SIGNAL_TOUCH_SENSOR = 1
FTR_SIGNAL_TAKE_OFF = 2
FTR_SIGNAL_FAKE_SOURCE = 3

# Response constants
FTR_CONTINUE = 0
FTR_CANCEL = 1

# State constants
FTR_STATE_FRAME_PROVIDED = 1
FTR_STATE_SIGNAL_PROVIDED = 2

# Định nghĩa structures
class FTR_DATA(ctypes.Structure):
    _fields_ = [
        ("dwSize", ctypes.c_ulong),
        ("pData", ctypes.POINTER(ctypes.c_ubyte))
    ]

class FTR_PROGRESS(ctypes.Structure):
    _fields_ = [
        ("dwSize", ctypes.c_ulong),
        ("dwCount", ctypes.c_ulong),
        ("bIsRepeated", ctypes.c_bool),
        ("dwTotal", ctypes.c_ulong)
    ]

class FTR_ENROLL_DATA(ctypes.Structure):
    _fields_ = [
        ("dwSize", ctypes.c_ulong),
        ("dwQuality", ctypes.c_ulong)
    ]

# Định nghĩa callback function prototype
FTR_CB_STATE_CONTROL = ctypes.CFUNCTYPE(None, ctypes.c_ulong)

def initialize_dll():
    """Khởi tạo DLL cho Futronic SDK"""
    try:
        dll_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "models")
        print(f"==> [DEBUG] DLL Directory: {dll_dir}")
        
        # Thêm đường dẫn DLL vào PATH
        if dll_dir not in os.environ['PATH']:
            os.environ['PATH'] = dll_dir + os.pathsep + os.environ['PATH']
        print(f"==> [DEBUG] Updated PATH: {os.environ['PATH']}")
        
        # Tải DLL chính
        dll_path = os.path.join(dll_dir, "ftrScanAPI.dll")
        if not os.path.exists(dll_path):
            print(f"==> [DEBUG] Không tìm thấy file {dll_path}")
            return None
            
        ftr_sdk = ctypes.CDLL(dll_path)
        print("==> [DEBUG] Đã tải ftrScanAPI.dll thành công")
        
        # Cấu hình các hàm có sẵn
        ftr_sdk.ftrScanOpenDevice.argtypes = [ctypes.c_int, ctypes.c_int]
        ftr_sdk.ftrScanOpenDevice.restype = ctypes.c_void_p
        
        ftr_sdk.ftrScanRollStart.argtypes = [ctypes.c_void_p]
        ftr_sdk.ftrScanRollStart.restype = ctypes.c_int
        
        ftr_sdk.ftrScanRollGetImage.argtypes = [ctypes.c_void_p, ctypes.c_void_p, ctypes.c_ulong]
        ftr_sdk.ftrScanRollGetImage.restype = ctypes.c_int
        
        ftr_sdk.ftrScanRollAbort.argtypes = [ctypes.c_void_p]
        ftr_sdk.ftrScanRollAbort.restype = ctypes.c_int
        
        ftr_sdk.ftrScanSetProperty.argtypes = [ctypes.c_void_p, ctypes.c_int, ctypes.c_void_p]
        ftr_sdk.ftrScanSetProperty.restype = ctypes.c_int
        
        ftr_sdk.ftrScanSetOptions.argtypes = [ctypes.c_void_p, ctypes.c_int]
        ftr_sdk.ftrScanSetOptions.restype = ctypes.c_int
        
        print("==> [DEBUG] Đã cấu hình các hàm ftrScanAPI thành công")
        return ftr_sdk
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi khởi tạo DLL: {str(e)}")
        return None

# Biến toàn cục cho SDK và cảm biến
ftr_sdk = None
fingerprint_sensor_initialized = False
current_frame_buffer = None
image_width = 0
image_height = 0
image_size = 0

def initialize_futronic_sdk():
    """Khởi tạo Futronic SDK"""
    global ftr_sdk, fingerprint_sensor_initialized, image_width, image_height, image_size
    
    try:
        print("==> [DEBUG] Bắt đầu khởi tạo Futronic SDK...")
        
        # Tải DLL
        ftr_sdk = initialize_dll()
        if not ftr_sdk:
            print("==> [DEBUG] Không thể tải DLL!")
            return False
            
        # Mở thiết bị với tham số mặc định
        handle = ftr_sdk.ftrScanOpenDevice(0, 0)
        if not handle:
            print("==> [DEBUG] Không thể mở thiết bị vân tay!")
            return False
            
        # Đợi thiết bị khởi động
        time.sleep(2)
        
        # Thiết lập kích thước ảnh mặc định
        image_width = 320
        image_height = 480
        image_size = image_width * image_height
        
        # Thiết lập các tùy chọn cơ bản
        try:
            # Thiết lập chế độ chụp ảnh
            capture_mode = ctypes.c_int(0)  # Chế độ mặc định
            result = ftr_sdk.ftrScanSetProperty(handle, 1, ctypes.byref(capture_mode))
            if result != 0:
                print(f"==> [DEBUG] Không thể thiết lập chế độ chụp ảnh! Mã lỗi: {result}")
            
            # Thiết lập chất lượng ảnh
            quality = ctypes.c_int(1)  # Chất lượng cao
            result = ftr_sdk.ftrScanSetProperty(handle, 2, ctypes.byref(quality))
            if result != 0:
                print(f"==> [DEBUG] Không thể thiết lập chất lượng ảnh! Mã lỗi: {result}")
                
        except Exception as e:
            print(f"==> [DEBUG] Lỗi khi thiết lập tùy chọn: {str(e)}")
            # Tiếp tục thực thi vì đây không phải lỗi nghiêm trọng
        
        fingerprint_sensor_initialized = True
        print("==> [DEBUG] Futronic SDK đã được khởi tạo thành công!")
        return True
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi khởi tạo SDK: {str(e)}")
        if 'handle' in locals():
            try:
                ftr_sdk.ftrScanRollAbort(handle)
            except:
                pass
        return False

def create_state_callback():
    """Tạo callback function để theo dõi trạng thái"""
    def state_callback(context, state_mask, progress_ptr, signal, bitmap_ptr):
        if state_mask & FTR_STATE_FRAME_PROVIDED:
            print("==> [DEBUG] Đã nhận frame từ thiết bị")
        if state_mask & FTR_STATE_SIGNAL_PROVIDED:
            if signal == FTR_SIGNAL_TOUCH_SENSOR:
                print("==> [DEBUG] Phát hiện vân tay trên cảm biến")
            elif signal == FTR_SIGNAL_TAKE_OFF:
                print("==> [DEBUG] Vân tay đã rời khỏi cảm biến")
            elif signal == FTR_SIGNAL_FAKE_SOURCE:
                print("==> [DEBUG] Phát hiện vân tay giả!")
        return FTR_CONTINUE
    return FTR_CB_STATE_CONTROL(state_callback)

def capture_fingerprint():
    """Chụp ảnh vân tay"""
    global ftr_sdk, image_width, image_height, image_size
    
    try:
        if not ftr_sdk:
            print("==> [DEBUG] SDK chưa được khởi tạo!")
            return None
            
        # Mở thiết bị
        handle = ftr_sdk.ftrScanOpenDevice(0, 0)
        if not handle:
            print("==> [DEBUG] Không thể mở thiết bị vân tay!")
            return None
            
        # Đợi thiết bị khởi động
        time.sleep(1)
        
        # Bắt đầu quá trình quét
        result = ftr_sdk.ftrScanRollStart(handle)
        if result != 0:
            print(f"==> [DEBUG] Không thể bắt đầu quét! Mã lỗi: {result}")
            ftr_sdk.ftrScanRollAbort(handle)
            return None
            
        # Tạo buffer cho ảnh
        image_buffer = (ctypes.c_ubyte * image_size)()
        
        # Lấy ảnh
        result = ftr_sdk.ftrScanRollGetImage(handle, image_buffer, image_size)
        if result != 0:
            print(f"==> [DEBUG] Không thể lấy ảnh! Mã lỗi: {result}")
            ftr_sdk.ftrScanRollAbort(handle)
            return None
            
        # Chuyển đổi buffer thành ảnh
        image_array = np.frombuffer(image_buffer, dtype=np.uint8)
        image = image_array.reshape((image_height, image_width))
        
        # Chuyển đổi ảnh thành base64
        _, buffer = cv2.imencode('.png', image)
        image_base64 = base64.b64encode(buffer).decode('utf-8')
        
        # Kết thúc quá trình quét
        ftr_sdk.ftrScanRollAbort(handle)
        return image_base64
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi chụp vân tay: {str(e)}")
        if 'handle' in locals():
            try:
                ftr_sdk.ftrScanRollAbort(handle)
            except:
                pass
        return None

def enroll_fingerprint():
    """Đăng ký vân tay"""
    global ftr_sdk, image_width, image_height, image_size
    
    try:
        if not ftr_sdk:
            print("==> [DEBUG] SDK chưa được khởi tạo!")
            return None, "SDK chưa được khởi tạo"
            
        # Mở thiết bị
        handle = ftr_sdk.ftrScanOpenDevice(0, 0)
        if not handle:
            print("==> [DEBUG] Không thể mở thiết bị vân tay!")
            return None, "Không thể mở thiết bị vân tay"
            
        # Đợi thiết bị khởi động
        time.sleep(1)
        
        # Kiểm tra vân tay
        is_present = ctypes.c_bool()
        result = ftr_sdk.ftrScanIsFingerPresent(handle, ctypes.byref(is_present))
        if result != FTR_RETCODE_OK:
            print(f"==> [DEBUG] Không thể kiểm tra vân tay! Mã lỗi: {result}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return None, f"Không thể kiểm tra vân tay, mã lỗi: {result}"
            
        if not is_present.value:
            print("==> [DEBUG] Không có vân tay trên cảm biến!")
            ftr_sdk.ftrScanCloseDevice(handle)
            return None, "Không có vân tay trên cảm biến"
            
        # Tạo callback
        callback = create_state_callback()
        
        # Tạo cấu trúc dữ liệu cho template
        template_data = FTR_DATA()
        template_data.dwSize = 0
        template_data.pData = None
        
        # Tạo cấu trúc dữ liệu cho quá trình đăng ký
        enroll_data = FTR_ENROLL_DATA()
        enroll_data.dwSize = ctypes.sizeof(FTR_ENROLL_DATA)
        enroll_data.dwQuality = 0
        
        # Thực hiện đăng ký
        result = ftr_sdk.ftrScanEnroll(handle, ctypes.byref(enroll_data), ctypes.byref(template_data), callback)
        if result != FTR_RETCODE_OK:
            print(f"==> [DEBUG] Không thể đăng ký vân tay! Mã lỗi: {result}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return None, f"Không thể đăng ký vân tay, mã lỗi: {result}"
            
        # Kiểm tra kích thước template
        if template_data.dwSize <= 0 or template_data.pData is None:
            print("==> [DEBUG] Template không hợp lệ!")
            ftr_sdk.ftrScanCloseDevice(handle)
            return None, "Template không hợp lệ"
            
        # Chuyển đổi template thành bytes an toàn
        try:
            # Tạo buffer với kích thước cố định
            buffer_size = min(template_data.dwSize, 1024 * 1024)  # Giới hạn 1MB
            buffer = (ctypes.c_ubyte * buffer_size)()
            
            # Copy dữ liệu vào buffer
            for i in range(buffer_size):
                buffer[i] = template_data.pData[i]
            
            # Chuyển đổi buffer thành bytes
            template_bytes = bytes(buffer)
            template_base64 = base64.b64encode(template_bytes).decode('utf-8')
            
        except Exception as e:
            print(f"==> [DEBUG] Lỗi khi chuyển đổi template: {str(e)}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return None, f"Lỗi khi chuyển đổi template: {str(e)}"
        
        ftr_sdk.ftrScanCloseDevice(handle)
        return template_base64, "Đăng ký vân tay thành công"
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi đăng ký vân tay: {str(e)}")
        if 'handle' in locals():
            try:
                ftr_sdk.ftrScanCloseDevice(handle)
            except:
                pass
        return None, f"Lỗi khi đăng ký vân tay: {str(e)}"

def verify_fingerprint(template_data):
    """Xác thực vân tay"""
    global ftr_sdk, image_width, image_height, image_size
    
    try:
        if not ftr_sdk:
            print("==> [DEBUG] SDK chưa được khởi tạo!")
            return False
            
        # Mở thiết bị
        handle = ftr_sdk.ftrScanOpenDevice(0, 0)
        if not handle:
            print("==> [DEBUG] Không thể mở thiết bị vân tay!")
            return False
            
        # Đợi thiết bị khởi động
        time.sleep(1)
        
        # Kiểm tra vân tay
        is_present = ctypes.c_bool()
        result = ftr_sdk.ftrScanIsFingerPresent(handle, ctypes.byref(is_present))
        if result != FTR_RETCODE_OK:
            print(f"==> [DEBUG] Không thể kiểm tra vân tay! Mã lỗi: {result}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return False
            
        if not is_present.value:
            print("==> [DEBUG] Không có vân tay trên cảm biến!")
            ftr_sdk.ftrScanCloseDevice(handle)
            return False
            
        # Tạo callback
        callback = create_state_callback()
        
        # Tạo cấu trúc dữ liệu cho template mới
        new_template = FTR_DATA()
        new_template.dwSize = 0
        new_template.pData = None
        
        # Tạo cấu trúc dữ liệu cho quá trình đăng ký
        enroll_data = FTR_ENROLL_DATA()
        enroll_data.dwSize = ctypes.sizeof(FTR_ENROLL_DATA)
        enroll_data.dwQuality = 0
        
        # Thực hiện đăng ký để lấy template mới
        result = ftr_sdk.ftrScanEnroll(handle, ctypes.byref(enroll_data), ctypes.byref(new_template), callback)
        if result != FTR_RETCODE_OK:
            print(f"==> [DEBUG] Không thể lấy template mới! Mã lỗi: {result}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return False
            
        # Chuyển đổi template đã lưu từ base64
        saved_template_bytes = base64.b64decode(template_data)
        saved_template = FTR_DATA()
        saved_template.dwSize = len(saved_template_bytes)
        saved_template.pData = (ctypes.c_ubyte * saved_template.dwSize).from_buffer_copy(saved_template_bytes)
        
        # So sánh template
        match_score = ctypes.c_ulong()
        result = ftr_sdk.ftrScanVerify(handle, ctypes.byref(saved_template), ctypes.byref(new_template), ctypes.byref(match_score))
        if result != FTR_RETCODE_OK:
            print(f"==> [DEBUG] Không thể so sánh template! Mã lỗi: {result}")
            ftr_sdk.ftrScanCloseDevice(handle)
            return False
            
        # Kiểm tra điểm số khớp
        match_threshold = 80  # Ngưỡng khớp 80%
        is_match = match_score.value >= match_threshold
        
        ftr_sdk.ftrScanCloseDevice(handle)
        return is_match
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi xác thực vân tay: {str(e)}")
        if 'handle' in locals():
            try:
                ftr_sdk.ftrScanCloseDevice(handle)
            except:
                pass
        return False

def check_sensor_status():
    """Kiểm tra trạng thái cảm biến vân tay"""
    global ftr_sdk
    
    try:
        if not ftr_sdk:
            print("==> [DEBUG] SDK chưa được khởi tạo!")
            return False, "SDK chưa được khởi tạo"
            
        # Mở thiết bị
        handle = ftr_sdk.ftrScanOpenDevice(0, 0)
        if not handle:
            print("==> [DEBUG] Không thể mở thiết bị vân tay!")
            return False, "Không thể mở thiết bị vân tay"
            
        # Thử bắt đầu quét
        result = ftr_sdk.ftrScanRollStart(handle)
        
        # Kết thúc quét
        ftr_sdk.ftrScanRollAbort(handle)
        
        if result != 0:
            print(f"==> [DEBUG] Không thể bắt đầu quét! Mã lỗi: {result}")
            return False, f"Không thể bắt đầu quét, mã lỗi: {result}"
            
        return True, "Cảm biến hoạt động bình thường"
        
    except Exception as e:
        print(f"==> [DEBUG] Lỗi khi kiểm tra trạng thái thiết bị: {str(e)}")
        if 'handle' in locals():
            try:
                ftr_sdk.ftrScanRollAbort(handle)
            except:
                pass
        return False, f"Lỗi khi kiểm tra thiết bị: {str(e)}"

def get_fingerprint_sensor():
    """Lấy trạng thái cảm biến hiện tại"""
    return fingerprint_sensor_initialized

def cleanup_sdk():
    """Dọn dẹp và đóng SDK"""
    global ftr_sdk, fingerprint_sensor_initialized
    
    try:
        if ftr_sdk and fingerprint_sensor_initialized:
            ftr_sdk.FTRTerminate()
            logger.info("✅ Đã đóng Futronic SDK")
        fingerprint_sensor_initialized = False
        ftr_sdk = None
    except Exception as e:
        logger.error(f"❌ Lỗi khi đóng SDK: {str(e)}")

# Django views
def check_fingerprint_sensor(request):
    """API endpoint để kiểm tra trạng thái cảm biến"""
    if request.method == "GET":
        try:
            if not get_fingerprint_sensor():
                # Thử khởi tạo lại
                if initialize_futronic_sdk():
                    sensor_ok, sensor_message = check_sensor_status()
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Không thể khởi tạo SDK',
                        'sensor_initialized': False
                    })
            else:
                sensor_ok, sensor_message = check_sensor_status()
                
            return JsonResponse({
                'success': sensor_ok,
                'message': sensor_message,
                'sensor_initialized': fingerprint_sensor_initialized,
                'image_size': image_size,
                'image_width': image_width,
                'image_height': image_height
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Lỗi kiểm tra cảm biến: {str(e)}',
                'sensor_initialized': False
            })
    
    return JsonResponse({'success': False, 'message': 'Phương thức không được hỗ trợ!'})

# Khởi tạo SDK khi module được load
print("==> Đang khởi tạo Futronic SDK...")
if initialize_futronic_sdk():
    print("==> ✅ Futronic SDK đã được khởi tạo thành công!")
else:
    print("==> ❌ Không thể khởi tạo Futronic SDK!")


@csrf_exempt
@login_required(login_url="login")
def register_fingerprint(request):
    if request.method == "POST":
        try:
            mssv = request.POST.get('mssv')
            name = request.POST.get('name')
            if not mssv or not name:
                return JsonResponse({'success': False, 'message': 'Vui lòng cung cấp đầy đủ MSSV và tên!'})

            # Kiểm tra xem SDK đã được khởi tạo chưa
            if not get_fingerprint_sensor():
                if not initialize_futronic_sdk():
                    return JsonResponse({'success': False, 'message': 'Không thể khởi tạo cảm biến vân tay!'})

            # Lấy dữ liệu từ cơ sở dữ liệu
            conn = connect_to_db()
            cursor = conn.cursor()
            
            # Kiểm tra xem sinh viên đã tồn tại chưa
            cursor.execute("SELECT name FROM thongtin WHERE mssv = ?", (mssv,))
            student = cursor.fetchone()
            
            if not student:
                # Thêm sinh viên mới nếu chưa tồn tại
                cursor.execute("INSERT INTO thongtin (mssv, name) VALUES (?, ?)", (mssv, name))
                conn.commit()
            
            # Đăng ký vân tay
            logger.info(f"🔍 Bắt đầu đăng ký vân tay cho {name} ({mssv})...")
            template_data, message = enroll_fingerprint()
            
            if template_data:
                # Chuyển đổi template thành chuỗi base64
                template_base64 = base64.b64encode(template_data).decode('utf-8')
                
                # Lưu vân tay vào database
                cursor.execute("UPDATE thongtin SET vantay = ? WHERE mssv = ?", (template_base64, mssv))
                conn.commit()
                
                # Tạo ảnh từ dữ liệu vân tay
                image_path = capture_fingerprint()
                if image_path:
                    with open(image_path, 'rb') as f:
                        image_data = f.read()
                        image_base64 = base64.b64encode(image_data).decode('utf-8')
                else:
                    image_base64 = None
                
                conn.close()
                return JsonResponse({
                    'success': True,
                    'message': 'Đăng ký vân tay thành công!',
                    'mssv': mssv,
                    'name': name,
                    'fingerprint_image': f'data:image/jpeg;base64,{image_base64}' if image_base64 else None
                })
            else:
                conn.close()
                return JsonResponse({'success': False, 'message': f'Không thể đăng ký vân tay: {message}'})
            
        except Exception as e:
            logger.error(f"Lỗi khi đăng ký vân tay: {e}")
            return JsonResponse({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Phương thức không được hỗ trợ!'})

@csrf_exempt
@login_required(login_url="login")
def diemdanh_vantay(request):
    if request.method == "POST":
        try:
            lophp = request.POST.get('lophp')
            tiet = request.POST.get('tiet')
            
            if not lophp or not tiet:
                return JsonResponse({'success': False, 'message': 'Vui lòng cung cấp lớp và tiết!'})
            
            # Kiểm tra xem SDK đã được khởi tạo chưa
            if not get_fingerprint_sensor():
                if not initialize_futronic_sdk():
                    return JsonResponse({'success': False, 'message': 'Không thể khởi tạo cảm biến vân tay!'})
            
            # Lấy danh sách vân tay đã đăng ký
            conn = connect_to_db()
            cursor = conn.cursor()
            
            # Lấy danh sách sinh viên trong lớp có vân tay
            cursor.execute("""
                SELECT thongtin.mssv, thongtin.name, thongtin.vantay 
                FROM thongtin 
                INNER JOIN dkmh ON dkmh.mssv = thongtin.mssv 
                WHERE dkmh.lophp = ? AND thongtin.vantay IS NOT NULL
            """, (lophp,))
            
            students = cursor.fetchall()
            if not students:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không có sinh viên nào có vân tay trong lớp này!'})
            
            # Lấy mã lịch
            magv = str(request.user.id)
            current_date = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("""
                SELECT malich 
                FROM lich 
                WHERE lophp = ? AND tiet = ? AND ngay = ? AND magv = ?
            """, (lophp, tiet, current_date, magv))
            
            lich_row = cursor.fetchone()
            if not lich_row:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không tìm thấy lịch học phù hợp!'})
            
            malich = lich_row[0]
            
            # Chụp ảnh vân tay
            logger.info("🔍 Bắt đầu nhận diện vân tay...")
            image_path = capture_fingerprint()
            
            if not image_path:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không thể chụp ảnh vân tay'})
            
            # Tạo template từ ảnh vân tay
            temp_template_data, temp_message = enroll_fingerprint()
            if not temp_template_data:
                conn.close()
                return JsonResponse({'success': False, 'message': f'Không thể tạo template vân tay: {temp_message}'})
            
            # So sánh với các vân tay đã đăng ký
            best_match = None
            best_score = 0
            
            for mssv, name, vantay_base64 in students:
                if not vantay_base64:
                    continue
                
                # Chuyển đổi từ base64 về bytes
                stored_template = base64.b64decode(vantay_base64)
                
                # Xác thực vân tay
                is_match, message = verify_fingerprint(stored_template)
                
                if is_match:
                    best_match = (mssv, name)
                    break
            
            if best_match:
                mssv, name = best_match
                
                # Kiểm tra xem sinh viên đã điểm danh chưa
                cursor.execute("""
                    SELECT COUNT(*) FROM diemdanh 
                    WHERE mssv = ? AND malich = ?
                """, (mssv, malich))
                
                if cursor.fetchone()[0] > 0:
                    conn.close()
                    return JsonResponse({
                        'success': False, 
                        'message': f'{name} ({mssv}) đã được điểm danh!'
                    })
                
                # Thêm điểm danh
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                cursor.execute("""
                    INSERT INTO diemdanh (mssv, time, malich)
                    VALUES (?, ?, ?)
                """, (mssv, current_time, malich))
                
                conn.commit()
                conn.close()
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Điểm danh thành công!',
                    'mssv': mssv,
                    'name': name
                })
            else:
                conn.close()
                return JsonResponse({'success': False, 'message': 'Không tìm thấy vân tay phù hợp!'})
            
        except Exception as e:
            logger.error(f"Lỗi khi điểm danh bằng vân tay: {e}")
            return JsonResponse({'success': False, 'message': f'Lỗi hệ thống: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Phương thức không được hỗ trợ!'})

@csrf_exempt
@login_required(login_url="login")
def them_sv(request):
    from .traning import train_new_student  

    if request.method == "POST":
        mssv = request.POST.get("mssv").upper()
        ten =" ".join(word.capitalize() for word in request.POST.get("name").split())
        lop = request.POST.get("lop").upper()
        images = request.FILES.getlist("images")

        if not mssv or not ten or not lop:
            return JsonResponse({"error": "Vui lòng điền đầy đủ thông tin"}, status=400)

        if not images:
            return JsonResponse({"error": "Vui lòng chọn ít nhất một ảnh"}, status=400)

        conn = connect_to_db()
        cursor = conn.cursor()
        
        try:
            # cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
            # if cursor.fetchone()[0] > 0:
            #     return JsonResponse({"error": f"MSSV {mssv} đã tồn tại trong hệ thống"}, status=400)

            success, message = train_new_student(mssv, ten, lop, images)
            if not success:
                return JsonResponse({"error": message}, status=400)

            return JsonResponse({"message": message})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": f"Lỗi hệ thống: {str(e)}"}, status=500)
        
        finally:
            cursor.close()
            conn.close()

    return JsonResponse({"error": "Phương thức không hợp lệ"}, status=405)

@csrf_exempt
@login_required(login_url="login")
def add_folder(request):
    from .traning import train_new_student  # Import từ traning.py

    if request.method == "POST":
        conn = connect_to_db()
        cursor = conn.cursor()

        added_students = []
        errors = []

        try:
            files = request.FILES.getlist('files[]')
            relative_paths = {key: value for key, value in request.POST.items() if key.startswith('relative_paths[')}
            if not files:
                return JsonResponse({"error": "Không có file nào được chọn"}, status=400)

            # Nhóm các file theo MSSV dựa trên relative_paths
            students_files = {}
            for i, file in enumerate(files):
                relative_path_key = f"relative_paths[{i}]"
                relative_path = relative_paths.get(relative_path_key, '')
                if not relative_path or relative_path.strip() == "":
                    errors.append(f"File {file.name} không có đường dẫn tương đối, vui lòng chọn thư mục")
                    continue
                path_parts = relative_path.split('/')
                if len(path_parts) < 2:
                    errors.append(f"Đường dẫn {relative_path} không đúng định dạng thư mục con")
                    continue
                folder_name = path_parts[1]
                if folder_name not in students_files:
                    students_files[folder_name] = []
                students_files[folder_name].append(file)

            if not students_files:
                return JsonResponse({"error": "Không tìm thấy thư mục hợp lệ", "errors": errors}, status=400)

            for folder_name, images in students_files.items():
                parts = folder_name.split("_")
                if len(parts) < 3:
                    errors.append(f"Tên thư mục '{folder_name}' không đúng định dạng (mssv_ten_lop)")
                    continue

                mssv = parts[0].upper()  
                ten = " ".join(word.capitalize() for word in parts[1].split()) 
                lop = "_".join(parts[2:]).upper()  
                ten_co_dau = ten

                cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
                if cursor.fetchone()[0] > 0:
                    errors.append(f"MSSV {mssv} đã tồn tại, bỏ qua")
                    continue

                success, message = train_new_student(mssv, ten_co_dau, lop, images)
                if not success:
                    errors.append(f"Training thất bại cho {mssv}: {message}")
                    continue

                added_students.append({"mssv": mssv, "name": ten_co_dau, "lop": lop})

            if not added_students:
                return JsonResponse({"error": "Không có sinh viên nào được thêm", "errors": errors}, status=400)

            return JsonResponse({
                "message": "Thêm sinh viên và cập nhật database thành công",
                "added": added_students,
                "errors": errors if errors else []
            })

        except Exception as e:
            print(f"Lỗi hệ thống: {str(e)}")
            return JsonResponse({"error": "Chỉ thêm được tối đa 20 sinh viên một lần, mỗi sinh viên 5 ảnh!"}, status=500)

        finally:
            cursor.close()
            conn.close()

    return JsonResponse({"error": "Yêu cầu không hợp lệ"}, status=400)

@login_required(login_url="login")
def get_weekly_schedule(request):
    start_date = request.GET.get('start_date')
    malich = request.GET.get('malich')
    magv = str(request.user.id)  # Lấy magv của người dùng hiện tại
    
    if not start_date and not malich:
        return JsonResponse({'status': 'fail', 'message': 'Thiếu ngày bắt đầu hoặc mã lịch'})
    
    try:
        conn = connect_to_db()
        cursor = conn.cursor()
        
        query = """
            SELECT malich, lophp, tiet, ngay 
            FROM lich 
            WHERE magv = ?
        """
        params = [magv]
        if malich:
            query += " AND malich = ?"
            params.append(str(malich))
        elif start_date:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = start + timedelta(days=6)
            query += " AND ngay BETWEEN ? AND ?"
            params.extend([start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')])
        query += " ORDER BY lophp, ngay, tiet"
        
        cursor.execute(query, params)
        schedule_data = cursor.fetchall()
        print(f"DEBUG: Dữ liệu lịch trong get_weekly_schedule cho magv {magv}: {schedule_data}")  # Debug
        
        if not schedule_data:
            return JsonResponse({'status': 'success', 'schedule': {}})

        weekly_schedule = {}
        for row in schedule_data:
            malich, lophp, tiet, ngay = row
            if lophp not in weekly_schedule:
                weekly_schedule[lophp] = {
                    'mon': [], 'tue': [], 'wed': [], 'thu': [],
                    'fri': [], 'sat': [], 'sun': []
                }
            day = datetime.strptime(ngay, '%Y-%m-%d').weekday()
            day_map = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
            weekly_schedule[lophp][day_map[day]].append({
                'malich': str(malich),
                'tiet': tiet,
                'ngay': ngay
            })
        
        conn.close()
        print(f"DEBUG: weekly_schedule trong get_weekly_schedule: {weekly_schedule}")  # Debug
        return JsonResponse({'status': 'success', 'schedule': weekly_schedule})
    
    except Exception as e:
        print(f"DEBUG: Lỗi trong get_weekly_schedule: {str(e)}")  # Debug
        return JsonResponse({'status': 'fail', 'message': str(e)})

from django import template
register = template.Library()

@register.filter
def get_item(dictionary, key):
    return dictionary.get(key, [])

@login_required(login_url="login")
@admin_required_sql
def lich_view(request):
    conn = connect_to_db()
    cursor = conn.cursor()
    
    magv = str(request.user.id)  
    print(f"DEBUG: magv của người dùng hiện tại: {magv}") 
    
    if request.method == "POST":
        action = request.POST.get('action')
        
        try:
            if action == 'add':
                cursor.execute("""
                    INSERT INTO lich (lophp, tiet, ngay, magv)
                    VALUES (?, ?, ?, ?)
                """, (
                    request.POST.get('lophp'),
                    request.POST.get('tiet'),
                    request.POST.get('ngay'),
                    magv,
                ))
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Thêm lịch thành công'})

            elif action == 'edit':
                cursor.execute("""
                    UPDATE lich SET lophp=?, tiet=?, ngay=?, magv=?
                    WHERE malich=? AND magv=?
                """, (
                    request.POST.get('lophp'),
                    request.POST.get('tiet'),
                    request.POST.get('ngay'),
                    magv,
                    request.POST.get('malich'),
                    magv,
                ))
                if cursor.rowcount == 0:
                    return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy lịch để sửa hoặc bạn không có quyền'})
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Sửa lịch thành công'})

            elif action == 'delete':
                malich = request.POST.get('malich')
                cursor.execute("DELETE FROM lich WHERE malich=? AND magv=?", (malich, magv))
                if cursor.rowcount == 0:
                    return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy lịch để xóa hoặc bạn không có quyền'})
                conn.commit()
                return JsonResponse({'status': 'success', 'message': 'Xóa lịch thành công'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})
        
        finally:
            conn.close()
    
    cursor.execute("SELECT DISTINCT lophp FROM lich WHERE magv = ?", (magv,))
    classes = [row[0] for row in cursor.fetchall()]
    
    weekly_schedule = {}
    for class_name in classes:
        weekly_schedule[class_name] = {
            'mon': [], 'tue': [], 'wed': [], 'thu': [],
            'fri': [], 'sat': [], 'sun': []
        }
        cursor.execute("""
            SELECT malich, lophp, tiet, ngay 
            FROM lich 
            WHERE lophp = ? AND magv = ?
            ORDER BY ngay, tiet
        """, (class_name, magv))
        schedule_data = cursor.fetchall()
        print(f"DEBUG: Lịch cho lớp {class_name} và magv {magv}: {schedule_data}")  
        
        for row in schedule_data:
            malich, lophp, tiet, ngay = row
            day = datetime.strptime(ngay, '%Y-%m-%d').weekday()
            day_map = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
            weekly_schedule[lophp][day_map[day]].append({
                'malich': str(malich),
                'tiet': tiet,
                'ngay': ngay
            })
    
    conn.close()
    return render(request, 'lich.html', {'weekly_schedule': weekly_schedule})


@login_required(login_url="login")
def get_members(request):
    lophp = request.GET.get('lophp')

    if not lophp:
        return JsonResponse({'status': 'fail', 'message': 'Vui lòng cung cấp lớp học phần'})

    conn = connect_to_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT DISTINCT thongtin.mssv, thongtin.name, thongtin.lop 
            FROM thongtin 
            INNER JOIN dkmh ON dkmh.mssv = thongtin.mssv 
            WHERE dkmh.lophp = ?
        """, (lophp,))
        students = cursor.fetchall()

        students_sorted = sorted(students, key=lambda x: vietnamese_last_word_sort_key(x[1]))

        data = {
            'status': 'success',
            'members': [
                {
                    'stt': idx + 1, 
                    'mssv': student[0],
                    'name': student[1],  
                    'lophp': student[2]
                }
                for idx, student in enumerate(students_sorted)
            ]
        }
        return JsonResponse(data)

    except Exception as e:
        print(f"Lỗi khi truy vấn dữ liệu thành viên: {e}")
        return JsonResponse({'status': 'fail', 'message': 'Lỗi khi lấy dữ liệu thành viên lớp'})

    finally:
        cursor.close()
        conn.close()

@csrf_exempt
@login_required(login_url="login")
def add_student_to_class(request):
    if request.method == "POST":
        mssv = request.POST.get('mssv').upper()
        lophp = request.POST.get('lophp')

        if not mssv or not lophp:
            return JsonResponse({'status': 'fail', 'message': 'Thiếu MSSV hoặc lớp học phần'})

        conn = connect_to_db()
        cursor = conn.cursor()

        try:
            cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
            if cursor.fetchone()[0] == 0:
                return JsonResponse({'status': 'fail', 'message': f'MSSV {mssv} không tồn tại trong hệ thống'})

            cursor.execute("SELECT COUNT(*) FROM dkmh WHERE mssv = ? AND lophp = ?", (mssv, lophp))
            if cursor.fetchone()[0] > 0:
                return JsonResponse({'status': 'fail', 'message': f'Sinh viên {mssv} đã đăng ký lớp {lophp}'})

            cursor.execute("INSERT INTO dkmh (mssv, lophp) VALUES (?, ?)", (mssv, lophp))
            conn.commit()
            return JsonResponse({'status': 'success', 'message': f'Thêm sinh viên {mssv} vào lớp {lophp} thành công'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'message': f'Lỗi hệ thống: {str(e)}'})
        
        finally:
            cursor.close()
            conn.close()
    return JsonResponse({'status': 'fail', 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
@login_required(login_url="login")
def update_student_in_class(request):
    if request.method == "POST":
        old_mssv = request.POST.get('old_mssv')
        new_mssv = request.POST.get('new_mssv')
        lophp = request.POST.get('lophp')

        if not old_mssv or not new_mssv or not lophp:
            return JsonResponse({'status': 'fail', 'message': 'Thiếu thông tin cần thiết'})

        conn = connect_to_db()
        cursor = conn.cursor()

        try:
            cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (new_mssv,))
            if cursor.fetchone()[0] == 0:
                return JsonResponse({'status': 'fail', 'message': 'MSSV mới không tồn tại trong hệ thống'})

            cursor.execute("UPDATE dkmh SET mssv = ? WHERE mssv = ? AND lophp = ?", (new_mssv, old_mssv, lophp))
            if cursor.rowcount == 0:
                return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy sinh viên trong lớp học phần'})
            conn.commit()
            return JsonResponse({'status': 'success', 'message': 'Sửa sinh viên thành công'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})
        
        finally:
            cursor.close()
            conn.close()
    return JsonResponse({'status': 'fail', 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
@login_required(login_url="login")
def delete_student_from_class(request):
    if request.method == "POST":
        mssv = request.POST.get('mssv')
        lophp = request.POST.get('lophp')

        if not mssv or not lophp:
            return JsonResponse({'status': 'fail', 'message': 'Thiếu MSSV hoặc lớp học phần'})

        conn = connect_to_db()
        cursor = conn.cursor()

        try:
            cursor.execute("DELETE FROM dkmh WHERE mssv = ? AND lophp = ?", (mssv, lophp))
            if cursor.rowcount == 0:
                return JsonResponse({'status': 'fail', 'message': 'Không tìm thấy sinh viên trong lớp học phần'})
            conn.commit()
            return JsonResponse({'status': 'success', 'message': 'Xóa sinh viên thành công'})

        except Exception as e:
            return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})
        
        finally:
            cursor.close()
            conn.close()
    return JsonResponse({'status': 'fail', 'message': 'Phương thức không hợp lệ'})

@csrf_exempt
@login_required(login_url="login")
def import_students_to_class_excel(request):
    if request.method != "POST" or not request.FILES.get('file') or not request.POST.get('lophp'):
        return JsonResponse({'status': 'fail', 'message': 'Thiếu lớp học phần hoặc file Excel'})

    lophp = request.POST.get('lophp').strip()
    excel_file = request.FILES['file']
    if not excel_file.name.endswith('.xlsx'):
        return JsonResponse({'status': 'fail', 'message': 'File phải có định dạng .xlsx'})

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file_path = temp_file.name
    
    try:
        with open(temp_file_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)
        
        wb = openpyxl.load_workbook(temp_file_path)
        sheet = wb.active
        if not sheet:
            raise ValueError("File Excel không chứa worksheet nào")
        
        conn = connect_to_db()
        cursor = conn.cursor()
        
        rows_added = 0
        errors = []
        expected_headers = ['MSSV']
        
        headers = [cell.value for cell in sheet[1]]
        if not all(h in headers for h in expected_headers):
            raise ValueError("File Excel không đúng định dạng. Cần cột: MSSV")
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 1:
                errors.append(f"Dòng {row_idx}: Thiếu dữ liệu")
                continue
            
            mssv = str(row[0]).strip() if row[0] else None
            if not mssv:
                errors.append(f"Dòng {row_idx}: MSSV không hợp lệ")
                continue
                
            cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
            if cursor.fetchone()[0] == 0:
                errors.append(f"Dòng {row_idx}: MSSV {mssv} không tồn tại trong hệ thống")
                continue
                
            cursor.execute("SELECT COUNT(*) FROM dkmh WHERE mssv = ? AND lophp = ?", (mssv, lophp))
            if cursor.fetchone()[0] > 0:
                errors.append(f"Dòng {row_idx}: MSSV {mssv} đã có trong lớp {lophp}")
                continue
                
            cursor.execute("INSERT INTO dkmh (mssv, lophp) VALUES (?, ?)", (mssv, lophp))
            rows_added += 1
        
        if rows_added == 0:
            raise ValueError("Không có dữ liệu hợp lệ nào được thêm")
        
        conn.commit()
        message = f"Nhập danh sách sinh viên thành công, đã thêm {rows_added} sinh viên vào lớp {lophp}"
        if errors:
            message += f". Có {len(errors)} lỗi: {', '.join(errors[:5])}"
        
        return JsonResponse({'status': 'success', 'message': message})
    
    except Exception as e:
        print(f"Lỗi khi nhập Excel: {e}")
        return JsonResponse({'status': 'fail', 'message': f'Lỗi: {str(e)}'})
    
    finally:
        cursor.close()
        conn.close()
        try:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
        except Exception as e:
            print(f"Không thể xóa file tạm: {e}")

@login_required(login_url="login")
def export_students_by_class(request):
    conn = connect_to_db()
    cursor = conn.cursor()

    lophp = request.GET.get('lophp')
    if not lophp:
        return HttpResponse("Vui lòng cung cấp lớp học phần trước khi xuất Excel", status=400)

    try:
        cursor.execute("""
            SELECT thongtin.mssv, thongtin.name, thongtin.lop 
            FROM thongtin 
            INNER JOIN dkmh ON dkmh.mssv = thongtin.mssv 
            WHERE dkmh.lophp = ?
            ORDER BY CAST(SUBSTR(thongtin.mssv, 2) AS INTEGER) ASC
        """, (lophp,))
        data = cursor.fetchall()

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="danhsach_{lophp}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Danh Sách {lophp}'

        columns = ['MSSV', 'Họ Tên', 'Lớp']
        ws.append(columns)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

        for row in data:
            ws.append([str(value) for value in row])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(response)
        return response

    except Exception as e:
        print(f"Lỗi khi xuất Excel: {e}")
        return HttpResponse(f"Lỗi: {str(e)}", status=500)

    finally:
        cursor.close()
        conn.close()
@csrf_exempt
@login_required(login_url="login")
def import_excel(request):
    if request.method != "POST" or not request.FILES.get('file'):
        return JsonResponse({'status': 'error', 'message': 'Không có file hoặc phương thức không hợp lệ'})

    excel_file = request.FILES['file']
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'status': 'error', 'message': 'File phải có định dạng .xlsx hoặc .xls'})

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xls' if excel_file.name.endswith('.xls') else '.xlsx')
    temp_file_path = temp_file.name
    
    conn = None
    cursor = None
    magv = str(request.user.id)
    print(magv)
    try:
        with open(temp_file_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)
        
        # Xử lý file dựa trên định dạng
        if excel_file.name.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(temp_file_path)
            sheet = wb.sheet_by_index(0)
            
            def get_cell_value(row_idx, col_idx):
                try:
                    return sheet.cell_value(row_idx - 1, col_idx - 1)
                except IndexError:
                    return None
            
            lophp = get_cell_value(8, 6)
            if not lophp or "Mã môn học/Mã nhóm:" not in lophp:
                raise ValueError("Không tìm thấy mã lớp học phần trong file Excel")
            lophp = lophp.split(":")[1].strip()
            lophp = lophp.replace("/", "-")

            headers = [get_cell_value(12, col) for col in range(1, sheet.ncols + 1)]
            date_columns = headers[11:]

            def iter_rows(min_row):
                for row_idx in range(min_row, sheet.nrows + 1):
                    yield [get_cell_value(row_idx, col) for col in range(1, sheet.ncols + 1)]
        
        elif excel_file.name.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(temp_file_path)
            sheet = wb.active
            if not sheet:
                raise ValueError("File Excel không chứa worksheet nào")
            
            lophp = sheet['F8'].value
            if not lophp or "Mã môn học/Mã nhóm:" not in lophp:
                raise ValueError("Không tìm thấy mã lớp học phần trong file Excel")
            lophp = lophp.split(":")[1].strip()

            headers = [cell.value for cell in sheet[12]]
            date_columns = headers[12:]

            def iter_rows(min_row):
                return sheet.iter_rows(min_row=min_row, values_only=True)

        conn = connect_to_db()
        cursor = conn.cursor()
        
        
        lich_success = []
        lich_failed = []
        sv_success = []
        sv_failed = []
        dd_success = 0
        
        lich_dict = {}
        for idx, date in enumerate(date_columns, start=1):
            if not date:
                continue
            try:
                date_obj = datetime.strptime(date, '%d/%m/%Y')
                date_str = date_obj.strftime('%Y-%m-%d')
            except ValueError:
                lich_failed.append(f"Ngày {date}: Định dạng không hợp lệ")
                continue

            cursor.execute(
                "SELECT malich FROM lich WHERE lophp = ? AND tiet = ? AND ngay = ? AND magv = ?",
                (lophp, "1", date_str, magv)
            )
            existing_lich = cursor.fetchone()
            if existing_lich:
                lich_dict[date] = existing_lich[0]
                lich_failed.append(f"Ngày {date}: Đã tồn tại")
            else:
                cursor.execute(
                    "INSERT INTO lich (lophp, tiet, ngay, magv) VALUES (?, ?, ?, ?)",
                    (lophp, "1", date_str, magv)
                )
                cursor.execute("SELECT last_insert_rowid()")
                malich = cursor.fetchone()[0]
                lich_dict[date] = malich
                lich_success.append(f"Ngày {date}")

        # Bước 2: Import danh sách sinh viên (bảng dkmh)
        for row_idx, row in enumerate(iter_rows(min_row=13), start=13):
            if not row or len(row) < 2:
                sv_failed.append(f"Dòng {row_idx}: Thiếu dữ liệu")
                continue

            mssv = str(row[1]).strip() if row[1] else None
            if not mssv:
                sv_failed.append(f"Dòng {row_idx}: MSSV không hợp lệ")
                continue

            cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
            if cursor.fetchone()[0] == 0:
                sv_failed.append(f"MSSV {mssv}: Không tồn tại trong hệ thống")
                continue

            cursor.execute(
                "SELECT COUNT(*) FROM dkmh WHERE lophp = ? AND mssv = ?",
                (lophp, mssv)
            )
            if cursor.fetchone()[0] > 0:
                sv_failed.append(f"MSSV {mssv}: Đã tồn tại trong lớp {lophp}")
                continue

            cursor.execute(
                "INSERT INTO dkmh (lophp, mssv) VALUES (?, ?)",
                (lophp, mssv)
            )
            sv_success.append(f"MSSV {mssv}")

        # Bước 3: Import điểm danh (bảng diemdanh)
        for row_idx, row in enumerate(iter_rows(min_row=13), start=13):
            if not row or len(row) < 2:
                continue

            mssv = str(row[1]).strip() if row[1] else None
            if not mssv:
                continue

            for col_idx, date in enumerate(date_columns, start=9):
                status = row[col_idx]
                if status == '✓':
                    date_str = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
                    if date in lich_dict:
                        malich = lich_dict[date]
                        cursor.execute(
                            "SELECT COUNT(*) FROM diemdanh WHERE mssv = ? AND malich = ?",
                            (mssv, malich)
                        )
                        if cursor.fetchone()[0] == 0:
                            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            cursor.execute(
                                "INSERT INTO diemdanh (mssv, time, malich) VALUES (?, ?, ?)",
                                (mssv, current_time, malich)
                            )
                            dd_success += 1

        conn.commit()
        
        message = f"Nhập Excel thành công:\n"
        message += f"- Thêm {len(lich_success)} lịch học thành công: {', '.join(lich_success[:5]) if lich_success else 'Không có'}\n"
        message += f"- Thêm {len(sv_success)} sinh viên thành công: {', '.join(sv_success[:5]) if sv_success else 'Không có'}\n"
        message += f"- Thêm {dd_success} bản ghi điểm danh\n"
        if lich_failed or sv_failed:
            message += "Các thất bại:\n"
            if lich_failed:
                message += f"- Lịch thất bại ({len(lich_failed)}): {', '.join(lich_failed[:5])}\n"
            if sv_failed:
                message += f"- Sinh viên thất bại ({len(sv_failed)}): {', '.join(sv_failed[:5])}"
        
        return JsonResponse({'status': 'success', 'message': message})
    
    except Exception as e:
        print(f"Lỗi khi nhập Excel: {e}")
        return JsonResponse({'status': 'error', 'message': f'Lỗi: {str(e)}'})
    
    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None:
            conn.close()
        try:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
        except Exception as e:
            print(f"Không thể xóa file tạm: {e}")

@login_required(login_url="login")
def export_excel(request):
    start_date = request.GET.get('start_date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = start + timedelta(days=6)
        
        conn = connect_to_db()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT malich, lophp, tiet, ngay FROM lich WHERE ngay BETWEEN ? AND ? ORDER BY ngay, lophp, tiet",
            (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))
        )
        lich_list = cursor.fetchall()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="lich_phong_{start.strftime("%Y%m%d")}_to_{end.strftime("%Y%m%d")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lịch Phòng'
        
        columns = ['maLich', 'maLop', 'tiet', 'ngay']
        ws.append(columns)
        
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
        
        for row in lich_list:
            ws.append([str(value) for value in row])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length
        
        wb.save(response)
        return response
    
    except Exception as e:
        print(f"Lỗi khi xuất Excel: {e}")
        return HttpResponse(f"Lỗi: {str(e)}", status=500)
    
    finally:
        cursor.close()
        conn.close()

@login_required(login_url="login")
def import_students_excel(request):
    if request.method != "POST" or not request.FILES.get('file'):
        return JsonResponse({'status': 'error', 'message': 'Không có file hoặc phương thức không hợp lệ'})

    excel_file = request.FILES['file']
    if not excel_file.name.endswith('.xlsx'):
        return JsonResponse({'status': 'error', 'message': 'File phải có định dạng .xlsx'})

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file_path = temp_file.name
    
    try:
        with open(temp_file_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)
        
        wb = openpyxl.load_workbook(temp_file_path)
        sheet = wb.active
        if not sheet:
            raise ValueError("File Excel không chứa worksheet nào")
        
        conn = connect_to_db()
        cursor = conn.cursor()
        
        rows_added = 0
        errors = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3:
                mssv = str(row[0]).strip() if row[0] else None
                name = str(row[1]).strip() if row[1] else None
                lop = str(row[2]).strip() if row[2] else None
                if mssv and name and lop:
                    cursor.execute("SELECT COUNT(*) FROM thongtin WHERE mssv = ?", (mssv,))
                    if cursor.fetchone()[0] > 0:
                        errors.append(f"MSSV {mssv} đã tồn tại, bỏ qua")
                        continue
                    
                    cursor.execute("INSERT INTO thongtin (mssv, name, lop) VALUES (?, ?, ?)", (mssv, name, lop))
                    rows_added += 1
        
        if rows_added == 0 and not errors:
            raise ValueError("Không có dữ liệu hợp lệ trong file Excel để thêm")
        
        conn.commit()
        conn.close()
        wb.close()
        
        message = f"Nhập Excel thành công, đã thêm {rows_added} sinh viên"
        if errors:
            message += f". Lỗi: {', '.join(errors)}"
        return JsonResponse({'status': 'success', 'message': message})
    
    except Exception as e:
        print(f"Lỗi khi nhập Excel sinh viên: {str(e)}")
        return JsonResponse({'status': 'error', 'message': f'Lỗi: {str(e)}'})
    
    finally:
        try:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
        except Exception as e:
            print(f"Không thể xóa file tạm: {e}")

@login_required(login_url="login")
def export_thongke_excel(request):
    lop = request.GET.get('lop', '').strip()
    if not lop:
        return HttpResponse("Vui lòng chọn lớp trước khi xuất Excel", status=400)

    magv = str(request.user.id)
    try:
        conn = connect_to_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT COUNT(*) 
            FROM lich 
            WHERE lophp = ? AND magv = ?
        """, (lop, magv))
        if cursor.fetchone()[0] == 0:
            conn.close()
            return HttpResponse("Bạn không có quyền xuất thống kê cho lớp này", status=403)

        cursor.execute("""
            SELECT malich, ngay, tiet 
            FROM lich 
            WHERE lophp = ? AND magv = ? 
            ORDER BY ngay, tiet
        """, (lop, magv))
        lich_rows = cursor.fetchall()

        headers = ['STT', 'MSSV', 'Họ Tên']
        lich_dict = {}
        for malich, ngay, tiet in lich_rows:
            headers.append(f"Tiết {tiet} Ngày {ngay}")
            lich_dict[str(malich)] = (ngay, tiet)

        cursor.execute("""
            SELECT DISTINCT thongtin.mssv, thongtin.name 
            FROM thongtin 
            INNER JOIN dkmh ON thongtin.mssv = dkmh.mssv 
            WHERE dkmh.lophp = ?
        """, (lop,))
        sv_rows = cursor.fetchall()

        sv_rows_sorted = sorted(sv_rows, key=lambda x: vietnamese_last_word_sort_key(x[1]))

        data = []
        for idx, (mssv, name) in enumerate(sv_rows_sorted, start=1):
            row_data = [idx, mssv, name]
            cursor.execute("""
                SELECT malich 
                FROM diemdanh 
                WHERE mssv = ?
            """, (mssv,))
            diemdanh_rows = cursor.fetchall()
            diemdanh_malich = set(row[0] for row in diemdanh_rows if row[0] is not None)

            for malich, (ngay, tiet) in lich_dict.items():
                row_data.append('✓' if malich in diemdanh_malich else '')
            data.append(row_data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="thongke_diemdanh_{lop}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Thống Kê Điểm Danh {lop}'

        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

        for row in data:
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                if cell.value == '✓':
                    cell.font = openpyxl.styles.Font(color='00FF00')
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

        wb.save(response)
        conn.close()
        return response

    except Exception as e:
        print(f"Lỗi trong export_thongke_excel: {e}")
        return HttpResponse(f"Lỗi: {str(e)}", status=500)

@login_required(login_url="login")
def thongke_chi_tiet(request):
    conn = connect_to_db()
    cursor = conn.cursor()
    if not is_superuser_sql(request.user.id):
        conn.close()
        return JsonResponse({'error': 'Không đủ quyền để truy cập!'}, status=403)
    lop = request.GET.get('lop', None)
    magv = str(request.user.id)

    try:
        cursor.execute("""
            SELECT DISTINCT lophp 
            FROM lich 
            WHERE magv = ? 
            ORDER BY lophp
        """, (magv,))
        lop_list = [row[0] for row in cursor.fetchall()]
        print(f"DEBUG: Danh sách lớp cho magv {magv}: {lop_list}")

        if not lop:
            return JsonResponse({
                'lop_list': lop_list,
                'tiet_ngay_stats': {'headers': [], 'data': []}
            })

        cursor.execute("""
            SELECT COUNT(*) 
            FROM lich 
            WHERE lophp = ? AND magv = ?
        """, (lop, magv))
        if cursor.fetchone()[0] == 0:
            return JsonResponse({
                'lop_list': lop_list,
                'tiet_ngay_stats': {'headers': [], 'data': []},
                'error': 'Bạn không có quyền xem thống kê lớp này'
            })

        cursor.execute("""
            SELECT malich, ngay, tiet 
            FROM lich 
            WHERE lophp = ? AND magv = ? 
            ORDER BY ngay, tiet
        """, (lop, magv))
        lich_rows = cursor.fetchall()
        
        headers = ['STT', 'MSSV', 'Họ Tên']
        lich_dict = {}
        for malich, ngay, tiet in lich_rows:
            headers.append(f"Tiết {tiet} Ngày {ngay}")
            lich_dict[str(malich)] = (ngay, tiet)

        cursor.execute("""
            SELECT DISTINCT thongtin.mssv, thongtin.name 
            FROM thongtin 
            INNER JOIN dkmh ON thongtin.mssv = dkmh.mssv 
            WHERE dkmh.lophp = ?
        """, (lop,))
        sv_rows = cursor.fetchall()

        # Sắp xếp theo tên tiếng Việt
        sv_rows_sorted = sorted(sv_rows, key=lambda x: vietnamese_last_word_sort_key(x[1]))

        data = []
        for idx, (mssv, name) in enumerate(sv_rows_sorted, start=1):
            row_data = {'stt': idx, 'mssv': mssv, 'name': name}
            cursor.execute("""
                SELECT malich 
                FROM diemdanh 
                WHERE mssv = ?
            """, (mssv,))
            diemdanh_rows = cursor.fetchall()
            diemdanh_malich = set(row[0] for row in diemdanh_rows if row[0] is not None)

            for malich, (ngay, tiet) in lich_dict.items():
                row_data[f"{ngay}_{tiet}"] = 'Đã điểm danh' if malich in diemdanh_malich else ''
            
            data.append(row_data)

        conn.close()
        return JsonResponse({
            'lop_list': lop_list,
            'tiet_ngay_stats': {
                'headers': headers,
                'data': data
            }
        })

    except Exception as e:
        print(f"Lỗi trong thongke_chi_tiet: {e}")
        return JsonResponse({'error': str(e)}, status=500)

def get_random_token(request):
    random_token = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
    request.session['random_token'] = random_token
    request.session.modified = True
    return JsonResponse({'random_token': random_token})


def register(request):
    if request.method == "POST":
        content_type = request.headers.get('Content-Type', '')
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                print("data", data)
                username = data.get('username')
                name = data.get('name')
                gmail = data.get('gmail')
                password = data.get('random_token')

                if not username or not name or not gmail or not password:
                    return JsonResponse({'success': False, 'message': 'Thiếu thông tin đăng ký!'})

                if User.objects.filter(username=username).exists():
                    return JsonResponse({'success': False, 'message': 'Tên đăng nhập đã tồn tại!'})
                elif User.objects.filter(email=gmail).exists():
                    return JsonResponse({'success': False, 'message': 'Gmail đã tồn tại!'})
                else:
                    user = User(username=username, email=gmail, last_name=name)
                    user.set_password(password)  
                    user.save()
                    return JsonResponse({'success': True, 'message': 'Đăng ký thành công! Hãy đăng nhập.'})
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': 'Dữ liệu không hợp lệ!'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không được hỗ trợ!'})

    return render(request, "register.html")

def login_view(request):
    if request.method == "POST":
        content_type = request.headers.get('Content-Type', '')
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                print("data", data)
                username = data.get('username')
                combined_data = data.get('random_token')
                if not username or not combined_data:
                    return JsonResponse({'success': False, 'message': 'Thiếu thông tin đăng nhập!'})
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': 'Dữ liệu không hợp lệ!'})
        else:
            return JsonResponse({'success': False, 'message': 'Yêu cầu không được hỗ trợ!'})

        user = authenticate(request, username=username, password=combined_data)
        if user:
            current_session_key = request.session.session_key

            all_sessions = Session.objects.filter(expire_date__gte=timezone.now())
            for session in all_sessions:
                session_data = session.get_decoded()
                session_user_id = session_data.get('_auth_user_id')
                if session_user_id == str(user.id) and session.session_key != current_session_key:
                    session.delete()

            login(request, user)
            request.session["user_id"] = user.id
            request.session["username"] = user.username
            request.session["name"] = user.last_name or user.username
            request.session.modified = True

            return JsonResponse({'success': True, 'message': 'Đăng nhập thành công!'})
        else:
            return JsonResponse({'success': False, 'message': 'Sai tên đăng nhập hoặc mật khẩu!'})

    return render(request, "login.html")
 

@login_required(login_url="login")
def update_profile(request):
    if request.method == "POST":
        user = request.user
        email = request.POST.get("email")
        name = request.POST.get("name")
        password = request.POST.get("hashed_password")
        print(email, name, password)
        try:
            if email and email != user.email:
                if User.objects.filter(email=email).exclude(id=user.id).exists():
                    messages.error(request, "Email đã được sử dụng!")
                    return redirect("update_profile")
                user.email = email

            if name:
                user.last_name = name

            if password:
                user.set_password(password)

            user.save()
            messages.success(request, "Cập nhật thông tin thành công!")
            return redirect("update_profile")

        except Exception as e:
            messages.error(request, f"Lỗi khi cập nhật: {str(e)}")
            return redirect("update_profile")
    return render(request, "profile.html")

@login_required
def check_session(request):
    is_authenticated = request.user.is_authenticated
    return JsonResponse({'is_authenticated': is_authenticated})

def logout_view(request):
    logout(request)
    request.session.flush()
    Session.objects.all().delete()
    return redirect('login')

def clear_all_sessions():
    try:
        logger.info("🔥 Server restart - Xóa tất cả session!")
        Session.objects.all().delete()
    except Exception as e:
        logger.error(f"Lỗi khi xóa session: {str(e)}")

# Khởi tạo TensorFlow và các model
try:
    import tensorflow as tf
    tf.get_logger().setLevel('ERROR')  # Giảm log level của TensorFlow
    
    # Khởi tạo model face detection
    mp_face_detection = mp.solutions.face_detection
    face_detection = mp_face_detection.FaceDetection(
        model_selection=0, 
        min_detection_confidence=0.5
    )
    
    # Khởi tạo model face recognition
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = InceptionResnetV1(pretrained='vggface2').eval().to(device)
    
    logger.info("✅ Đã khởi tạo các model thành công")
except Exception as e:
    logger.error(f"❌ Lỗi khi khởi tạo model: {str(e)}")



# Xóa session khi khởi động
clear_all_sessions()

def manual_open(request):
    if request.method == "POST":
        mssv = request.POST.get("manual_mssv")
        username = request.POST.get("username")
        password = request.POST.get("password")

        if not mssv or not username or not password:
            return JsonResponse({"error": "Thiếu thông tin đăng nhập hoặc MSSV!"}, status=400)

        mssv = mssv.strip()
        user = authenticate(username=username, password=password)
        if user is None:
            return JsonResponse({"error": "Tài khoản hoặc mật khẩu không chính xác!"}, status=400)

        conn = connect_to_db()
        cursor = conn.cursor()

        cursor.execute("SELECT mssv FROM thongtin WHERE mssv = ?", (mssv,))
        sinh_vien = cursor.fetchone()

        if not sinh_vien:
            conn.close()
            return JsonResponse({"error": "MSSV không tồn tại trong hệ thống!"}, status=400)

        time_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("INSERT INTO diemdanh (mssv, time) VALUES (?, ?)", (mssv, time_now))

        conn.commit()
        conn.close()

        return JsonResponse({"message": "✅ Đã log thành công!"})

    return JsonResponse({"error": "Yêu cầu không hợp lệ!"}, status=400)

def initialize_fingerprint_sensor():
    """Khởi tạo cảm biến vân tay sử dụng ftrScanAPI"""
    global ftr_scan, handle, fingerprint_sensor_initialized, image_width, image_height, image_size
    
    try:
        logger.info("🔍 Bắt đầu khởi tạo cảm biến vân tay FS80...")
        print("==> [DEBUG] Bắt đầu khởi tạo cảm biến vân tay FS80...")

        if not ftr_scan:
            logger.error("❌ DLL chưa được khởi tạo thành công!")
            print("==> [DEBUG] DLL chưa được khởi tạo thành công!")
            return None

        # Đóng thiết bị nếu đã mở
        try:
            if 'handle' in globals() and handle:
                if hasattr(ftr_scan, 'ftrScanCloseDevice'):
                    ftr_scan.ftrScanCloseDevice(handle)
                    print("==> [DEBUG] Đóng thiết bị đang mở")
        except Exception as close_error:
            print(f"==> [DEBUG] Lỗi khi đóng thiết bị: {str(close_error)}")
            pass

        # Đợi một chút trước khi mở thiết bị
        time.sleep(2)
        print("==> [DEBUG] Đợi 2 giây trước khi mở thiết bị")

        # Mở thiết bị
        try:
            print("==> [DEBUG] Thử mở thiết bị với ftrScanOpenDevice")
            if hasattr(ftr_scan, 'ftrScanOpenDevice'):
                handle = ftr_scan.ftrScanOpenDevice(0, 0)
                print(f"==> [DEBUG] ftrScanOpenDevice trả về: {handle}")
                
                if not handle or handle == 0:
                    logger.error("❌ Không thể mở thiết bị vân tay!")
                    print("==> [DEBUG] Không thể mở thiết bị vân tay!")
                    return None
            else:
                logger.error("❌ Không tìm thấy hàm ftrScanOpenDevice trong DLL!")
                print("==> [DEBUG] Không tìm thấy hàm ftrScanOpenDevice trong DLL!")
                return None
        except Exception as open_error:
            logger.error(f"❌ Lỗi khi mở thiết bị: {str(open_error)}")
            print(f"==> [DEBUG] Lỗi khi mở thiết bị: {str(open_error)}")
            return None

        # Đợi sau khi mở thiết bị
        time.sleep(3)
        print("==> [DEBUG] Đợi 3 giây sau khi mở thiết bị")

        # Lấy thông tin thiết bị
        try:
            if hasattr(ftr_scan, 'ftrScanGetDeviceInfo'):
                device_info = (ctypes.c_char * 100)()
                result = ftr_scan.ftrScanGetDeviceInfo(handle, device_info, 100)
                print(f"==> [DEBUG] ftrScanGetDeviceInfo trả về: {result}")
                
                if result == 0:  # 0 = success
                    info_str = device_info.value.decode('utf-8', errors='ignore')
                    print(f"==> [DEBUG] Thông tin thiết bị: {info_str}")
                    logger.info(f"✅ Thông tin thiết bị: {info_str}")
                else:
                    print(f"==> [DEBUG] Không thể lấy thông tin thiết bị, mã lỗi: {result}")
                    # Thử nhiều lần
                    max_attempts = 5
                    for attempt in range(max_attempts):
                        print(f"==> [DEBUG] Thử lại lần {attempt+1}/{max_attempts}")
                        time.sleep(1)
                        result = ftr_scan.ftrScanGetDeviceInfo(handle, device_info, 100)
                        if result == 0:
                            info_str = device_info.value.decode('utf-8', errors='ignore')
                            print(f"==> [DEBUG] Thông tin thiết bị: {info_str}")
                            logger.info(f"✅ Thông tin thiết bị: {info_str}")
                            break
                    if result != 0:
                        logger.error("❌ Không thể lấy thông tin thiết bị sau nhiều lần thử!")
                        print("==> [DEBUG] Không thể lấy thông tin thiết bị sau nhiều lần thử!")
            else:
                print("==> [DEBUG] Không tìm thấy hàm ftrScanGetDeviceInfo trong DLL")
        except Exception as info_error:
            logger.error(f"❌ Lỗi khi lấy thông tin thiết bị: {str(info_error)}")
            print(f"==> [DEBUG] Lỗi khi lấy thông tin thiết bị: {str(info_error)}")

        # Lấy kích thước ảnh
        try:
            if hasattr(ftr_scan, 'ftrScanGetImageSize'):
                image_size = ctypes.c_ulong()
                result = ftr_scan.ftrScanGetImageSize(handle, ctypes.byref(image_size))
                print(f"==> [DEBUG] ftrScanGetImageSize trả về: {result}, size: {image_size.value}")
                
                if result == 0:  # 0 = success
                    logger.info(f"✅ Kích thước ảnh: {image_size.value} bytes")
                else:
                    logger.error(f"❌ Không thể lấy kích thước ảnh, mã lỗi: {result}")
                    print(f"==> [DEBUG] Không thể lấy kích thước ảnh, mã lỗi: {result}")
            else:
                print("==> [DEBUG] Không tìm thấy hàm ftrScanGetImageSize trong DLL")
        except Exception as size_error:
            logger.error(f"❌ Lỗi khi lấy kích thước ảnh: {str(size_error)}")
            print(f"==> [DEBUG] Lỗi khi lấy kích thước ảnh: {str(size_error)}")

        # Kiểm tra xem vân tay có hiện diện hay không
        try:
            if hasattr(ftr_scan, 'ftrScanIsFingerPresent'):
                is_present = ctypes.c_bool()
                result = ftr_scan.ftrScanIsFingerPresent(handle, ctypes.byref(is_present))
                print(f"==> [DEBUG] ftrScanIsFingerPresent trả về: {result}, is_present: {is_present.value}")
                
                if result == 0:  # 0 = success
                    if is_present.value:
                        print("==> [DEBUG] Đã phát hiện vân tay trên cảm biến")
                    else:
                        print("==> [DEBUG] Không có vân tay trên cảm biến")
                else:
                    print(f"==> [DEBUG] Lỗi khi kiểm tra vân tay, mã lỗi: {result}")
            else:
                print("==> [DEBUG] Không tìm thấy hàm ftrScanIsFingerPresent trong DLL")
        except Exception as finger_error:
            print(f"==> [DEBUG] Lỗi khi kiểm tra vân tay: {str(finger_error)}")

        fingerprint_sensor_initialized = True
        logger.info("✅ Cảm biến vân tay đã được khởi tạo thành công!")
        print("==> [DEBUG] Cảm biến vân tay đã được khởi tạo thành công!")
        return handle
        
    except Exception as e:
        logger.error(f"❌ Lỗi khi khởi tạo cảm biến vân tay: {str(e)}")
        print(f"==> [DEBUG] Lỗi khi khởi tạo cảm biến vân tay: {str(e)}")
        return None

def cleanup_futronic_sdk():
    """Dọn dẹp và giải phóng tài nguyên SDK"""
    global ftr_sdk, fingerprint_sensor_initialized
    
    try:
        if ftr_sdk:
            # Mở thiết bị
            handle = ftr_sdk.ftrScanOpenDevice(0, 0)
            if handle:
                # Đóng thiết bị
                ftr_sdk.ftrScanCloseDevice(handle)
                logger.info("✅ Đã đóng thiết bị vân tay")
                print("==> [DEBUG] Đã đóng thiết bị vân tay")
            
            # Giải phóng DLL
            ftr_sdk = None
            fingerprint_sensor_initialized = False
            logger.info("✅ Đã giải phóng SDK")
            print("==> [DEBUG] Đã giải phóng SDK")
            
    except Exception as e:
        logger.error(f"❌ Lỗi khi dọn dẹp SDK: {str(e)}")
        print(f"==> [DEBUG] Lỗi khi dọn dẹp SDK: {str(e)}")



